# %%
import re
import statistics
import openpyxl
from openpyxl.styles import PatternFill
from dateutil import parser
from openpyxl import load_workbook


# -----------------------------
# Function 1: Highlight Outliers
# -----------------------------
def parse_value(val):
    """Convert cell value into numeric for comparison, keep None if invalid."""
    if val is None:
        return None
    val = str(val).strip()
    if val.lower() in ["na", "nan", "****", ""]:
        return None
    if re.match(r"^<\d*\.?\d+$", val):  # "<0.002"
        try:
            return float(val[1:])
        except:
            return None
    try:
        return float(val)
    except:
        return None


def is_less_than(val):
    """Check if value is in <number> format."""
    if val is None:
        return False
    return bool(re.match(r"^<\d*\.?\d+$", str(val).strip()))


def highlight_outliers(file_path):
    wb = openpyxl.load_workbook(file_path)
    yellow_fill = PatternFill(
        start_color="FFFF99", end_color="FFFF99", fill_type="solid"
    )

    unit_keywords = ["mg/L", "ug/L", "pH", "uS", "deg.C"]

    for sheet in wb.worksheets:
        if sheet.sheet_state != "visible":
            continue

        max_row = sheet.max_row
        max_col = sheet.max_column

        for row in range(3, max_row + 1):
            variable = (
                str(sheet.cell(row=row, column=2).value).strip()
                if sheet.cell(row=row, column=2).value
                else ""
            )
            if variable == "":
                continue

            raw_values = []
            col_map = {}
            for col in range(3, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                val = str(cell.value).strip() if cell.value else ""

                # skip unit columns
                if any(u in val for u in unit_keywords):
                    continue

                if val not in ["", "na", "NA", "NaN", "****"]:
                    raw_values.append(val)
                    col_map[col] = val

            if len(raw_values) < 3:
                continue  # too few points to judge outliers

            # --- Step 1: Pattern consistency (<number> vs plain number) ---
            less_than_flags = [is_less_than(v) for v in raw_values]

            if any(less_than_flags) and not all(less_than_flags):
                # Mixed pattern → highlight inconsistent ones
                majority = max(set(less_than_flags), key=less_than_flags.count)
                for col, raw_val in col_map.items():
                    if is_less_than(raw_val) != majority:
                        sheet.cell(row=row, column=col).fill = yellow_fill
                continue  # skip numeric outlier detection for this row

            # --- Step 2: Numeric outlier detection ---
            values = [parse_value(v) for v in raw_values if parse_value(v) is not None]
            if len(values) < 3:
                continue

            median_val = statistics.median(values)
            mad = statistics.median([abs(v - median_val) for v in values]) or 1e-6

            for col, raw_val in col_map.items():
                num = parse_value(raw_val)
                if num is None:
                    continue
                # robust outlier detection
                if abs(num - median_val) / mad > 6:  # threshold can be tuned
                    sheet.cell(row=row, column=col).fill = yellow_fill

    wb.save(file_path)
    print(f"Outliers highlighted in {file_path}")


# --------------------------------
# Function 2: Highlight Excel Errors
# --------------------------------
def highlight_excel_errors(file_path):
    wb = openpyxl.load_workbook(file_path)
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    unit_keywords = ["mg/L", "ug/L", "pH", "uS", "deg.C"]

    for sheet in wb.worksheets:
        if sheet.sheet_state != "visible":
            continue  # skip hidden sheets

        max_row = sheet.max_row
        max_col = sheet.max_column

        # Start from row 3 and col 3 (C)
        for row in range(3, max_row + 1):
            for col in range(3, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                val = str(cell.value).strip() if cell.value is not None else ""

                # Skip empty, na, or ****
                if val.lower() in ["na", "nan"] or val == "****" or val == "":
                    continue

                # Skip unit columns
                col_values = [
                    str(sheet.cell(r, col).value).strip()
                    for r in range(3, max_row + 1)
                    if sheet.cell(r, col).value
                ]
                if col_values and all(
                    any(u in v for u in unit_keywords) for v in col_values
                ):
                    continue

                # (a) Comma instead of decimal
                if "," in val and any(ch.isdigit() for ch in val):
                    cell.fill = red_fill
                    continue

                # (b) 'I' instead of '1'
                if "I" in val and any(ch.isdigit() for ch in val.replace("I", "1")):
                    cell.fill = red_fill
                    continue

                # (c) 'l' instead of '1'
                if "l" in val and any(ch.isdigit() for ch in val.replace("l", "1")):
                    cell.fill = red_fill
                    continue

                # (d) Space inside numbers (e.g., "0. 002", "0 008", "7 368")
                if re.search(r"\d\s+\d", val):
                    cell.fill = red_fill
                    continue

                # (e) 'o' instead of '0'
                if "o" in val:
                    cell.fill = red_fill
                    continue

                # (f) "<.008" type
                if val.startswith("<.") and len(val) > 2:
                    cell.fill = red_fill
                    continue

    wb.save(file_path)
    print(f"Errors highlighted in {file_path}")


# ----------------------------
# Function 3: Fix Date Formats
# ----------------------------
def convert_to_mdy(value):
    """Convert value to MM/DD/YYYY string if possible, else return original as string."""
    if value is None:
        return value
    try:
        dt = parser.parse(str(value), dayfirst=False, yearfirst=False)
        return dt.strftime("%m/%d/%Y")
    except:
        return str(value)  # always return as string


def fix_dates_in_excel(file_path, rows_with_dates=None, cols_with_dates=None):
    """
    file_path: path to the Excel file
    rows_with_dates: list of row numbers (1-based, like Excel)
    cols_with_dates: list of column letters (e.g., ["B","C"]) or numbers (1-based)
    """
    wb = load_workbook(file_path)

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]

        # If no rows/cols provided, default to all
        row_list = rows_with_dates if rows_with_dates else range(1, ws.max_row + 1)
        col_list = cols_with_dates if cols_with_dates else range(1, ws.max_column + 1)

        for row_num in row_list:
            if row_num <= ws.max_row:
                for col in col_list:
                    if isinstance(col, str):  # column as letter
                        cell = ws[f"{col}{row_num}"]
                    else:  # column as number
                        cell = ws.cell(row=row_num, column=col)

                    val = cell.value
                    if val is not None:
                        new_val = convert_to_mdy(val)
                        cell.value = str(new_val)  # force string
                        cell.number_format = "@"  # mark as Text

    new_path = file_path.replace(".xlsx", "_corrected.xlsx")
    wb.save(new_path)
    print(f"Processed and saved: {new_path}")


# %%
# ----------------------------
# ----------------------------
if __name__ == "__main__":
    file_path = r"D:\Projects\Data Extraction 8th August\Test\Analytical Data -Soil (BH)- 1999-10-20-REP-Stage 1 Preliminary Site Investigation_Lot F.xlsx"
    highlight_excel_errors(file_path)
    highlight_outliers(file_path)
    fix_dates_in_excel(file_path)
    pass

# %%
